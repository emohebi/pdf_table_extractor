"""
Invoice Reconciliation Tool

Compares extracted invoice data (JSON) against SAP records (Excel)
by matching on SES Number.

Algorithm:
    For each JSON file in the input directory:
        1. Extract "SES Number" from the JSON (uses LLM if not a top-level field)
        2. Extract "Document Total" from the JSON (uses LLM to locate it)
        3. Look up matching rows in the Excel file by SES Number
        4. Sum "Total Spend" for all matching rows -> "SAP Price"
        5. Store SES Number, SAP Price, Invoice Price

    Output: An Excel file with columns [SES Number, SAP Price, Invoice Price]

Usage:
    # Command line
    python -m src.utils.reconciler \\
        --json-dir ./output/intermediate \\
        --excel ./input/sap_data.xlsx \\
        --output ./output/reconciliation.xlsx

    # Python API
    from src.utils.reconciler import Reconciler

    reconciler = Reconciler(
        json_dir="./output/intermediate",
        excel_path="./input/sap_data.xlsx",
    )
    reconciler.run(output_path="./output/reconciliation.xlsx")
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.utils.logger import get_logger
from src.utils.file_utils import ensure_directory

logger = get_logger(__name__)


# ---------------------------------------------------------------------------
# LLM helpers
# ---------------------------------------------------------------------------

def _build_llm_prompt_ses(json_content: str) -> str:
    """Build a prompt to extract the SES Number from a JSON string."""
    return (
        "You are given the JSON content of an extracted invoice/contract table.\n"
        "Find the SES Number (also called Service Entry Sheet number).\n"
        "It is usually a numeric string (e.g. '5100012345').\n\n"
        "Return ONLY a JSON object with one key:\n"
        '{"ses_number": "<value or null>"}\n\n'
        "No markdown, no explanation.\n\n"
        f"JSON content (truncated to first 6000 chars):\n{json_content[:6000]}"
    )


def _build_llm_prompt_doc_total(json_content: str) -> str:
    """Build a prompt to extract the Document Total from a JSON string."""
    return (
        "You are given the JSON content of an extracted invoice/contract table.\n"
        "Find the Document Total (the overall total amount of the invoice).\n"
        "It might be labelled 'Document Total', 'Grand Total', 'Invoice Total', "
        "'Total Amount', 'Total', or similar.\n\n"
        "Return ONLY a JSON object with one key:\n"
        '{"document_total": <numeric value or null>}\n\n'
        "No markdown, no explanation.\n\n"
        f"JSON content (truncated to first 6000 chars):\n{json_content[:6000]}"
    )


def _call_llm(prompt: str, extractor) -> dict:
    """
    Call Azure OpenAI via the existing GPT4VisionExtractor client.

    Args:
        prompt: The text prompt to send.
        extractor: A GPT4VisionExtractor instance (we reuse its client).

    Returns:
        Parsed JSON dict from the model response.
    """
    response = extractor.client.chat.completions.create(
        model=extractor.deployment,
        messages=[{"role": "user", "content": prompt}],
        max_completion_tokens=256,
        temperature=0,
    )
    text = response.choices[0].message.content.strip()

    # Strip markdown fences if present
    if text.startswith("```"):
        lines = text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        logger.error(f"LLM returned unparseable response: {text[:300]}")
        return {}


# ---------------------------------------------------------------------------
# Core extraction helpers
# ---------------------------------------------------------------------------

def _try_top_level_ses(data: dict) -> Optional[str]:
    """Try to find SES Number as a top-level or shallow field."""
    # Check common key names (case-insensitive)
    target_keys = ["ses_number", "ses number", "sesnumber", "ses_no", "ses no"]

    def _search(obj, depth=0):
        if depth > 3:
            return None
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k.lower().replace("_", " ").strip() in target_keys:
                    return str(v) if v is not None else None
                found = _search(v, depth + 1)
                if found:
                    return found
        elif isinstance(obj, list):
            for item in obj:
                found = _search(item, depth + 1)
                if found:
                    return found
        return None

    return _search(data)


def _try_top_level_doc_total(data: dict) -> Optional[float]:
    """Try to find Document Total as a top-level or shallow field."""
    target_keys = [
        "document_total", "document total", "documenttotal",
        "grand_total", "grand total", "grandtotal",
        "invoice_total", "invoice total", "invoicetotal",
        "total_amount", "total amount", "totalamount",
        "total",
    ]

    def _search(obj, depth=0):
        if depth > 3:
            return None
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k.lower().replace("_", " ").strip() in target_keys:
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        return None
                found = _search(v, depth + 1)
                if found is not None:
                    return found
        elif isinstance(obj, list):
            for item in obj:
                found = _search(item, depth + 1)
                if found is not None:
                    return found
        return None

    return _search(data)


def extract_ses_number(data: dict, raw_json: str, extractor) -> Optional[str]:
    """
    Extract SES Number from JSON data.
    First tries a deterministic search; falls back to LLM.
    """
    ses = _try_top_level_ses(data)
    if ses:
        logger.info(f"  SES Number found deterministically: {ses}")
        return ses

    logger.info("  SES Number not found in keys, calling LLM...")
    result = _call_llm(_build_llm_prompt_ses(raw_json), extractor)
    ses = result.get("ses_number")
    if ses:
        logger.info(f"  SES Number from LLM: {ses}")
    else:
        logger.warning("  LLM could not find SES Number")
    return str(ses) if ses else None


def extract_document_total(data: dict, raw_json: str, extractor) -> Optional[float]:
    """
    Extract Document Total from JSON data.
    First tries a deterministic search; falls back to LLM.
    """
    total = _try_top_level_doc_total(data)
    if total is not None:
        logger.info(f"  Document Total found deterministically: {total}")
        return total

    logger.info("  Document Total not found in keys, calling LLM...")
    result = _call_llm(_build_llm_prompt_doc_total(raw_json), extractor)
    doc_total = result.get("document_total")
    if doc_total is not None:
        try:
            doc_total = float(doc_total)
            logger.info(f"  Document Total from LLM: {doc_total}")
            return doc_total
        except (TypeError, ValueError):
            pass
    logger.warning("  LLM could not find Document Total")
    return None


# ---------------------------------------------------------------------------
# Excel lookup
# ---------------------------------------------------------------------------

def lookup_sap_price(df: pd.DataFrame, ses_number: str) -> float:
    """
    Find all rows in the Excel DataFrame where SES Number matches
    and sum their Total Spend values.
    """
    # Normalise the SES column to string for comparison
    ses_col = df.columns[df.columns.str.lower().str.replace(" ", "").str.contains("sesnumber|sesno")][0]
    spend_col = df.columns[df.columns.str.lower().str.replace(" ", "").str.contains("totalspend(aud)", regex=False)][0]

    # Normalize both sides to clean integer strings.
    # Excel column is float64 with nulls (e.g. 5100012345.0, NaN),
    # so we drop nulls, convert to int, then to str for comparison.
    def _normalize_ses(val) -> Optional[str]:
        """Convert a SES value to a clean integer string, or None."""
        if pd.isna(val):
            return None
        try:
            return str(int(float(val)))
        except (ValueError, TypeError):
            # Already a string, just strip whitespace and trailing .0
            s = str(val).strip()
            if s.endswith(".0"):
                s = s[:-2]
            return s

    needle = _normalize_ses(ses_number)
    if needle is None:
        logger.warning(f"  SES Number '{ses_number}' could not be normalized")
        return 0.0

    normalized_col = df[ses_col].map(_normalize_ses)
    mask = normalized_col == needle
    matched = df.loc[mask, spend_col]

    total = matched.sum()
    logger.info(f"  SAP lookup: {len(matched)} rows matched, Total Spend sum = {total}")
    return float(total)


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_output_excel(records: list[dict], output_path: Union[str, Path]) -> Path:
    """
    Write reconciliation results to a formatted Excel file.
    """
    output_path = Path(output_path)
    ensure_directory(output_path.parent)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    currency_fmt = '#,##0.00'

    headers = ["SES Number", "SAP Price", "Invoice Price", "Difference", "Source File"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, rec in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=rec["ses_number"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=rec["sap_price"]).border = thin_border
        ws.cell(row=row_idx, column=2).number_format = currency_fmt
        ws.cell(row=row_idx, column=3, value=rec["invoice_price"]).border = thin_border
        ws.cell(row=row_idx, column=3).number_format = currency_fmt

        # Difference formula
        ws.cell(row=row_idx, column=4).border = thin_border
        ws.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}")
        ws.cell(row=row_idx, column=4).number_format = currency_fmt

        ws.cell(row=row_idx, column=5, value=rec.get("source_file", "")).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 40

    wb.save(str(output_path))
    logger.info(f"Output saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Main Reconciler class
# ---------------------------------------------------------------------------

class Reconciler:
    """
    Reconciles invoice JSON outputs against SAP Excel records.

    Example:
        from src.utils.reconciler import Reconciler

        r = Reconciler(
            json_dir="./output/intermediate",
            excel_path="./input/sap_data.xlsx",
            api_key="...",
            endpoint="https://...",
        )
        r.run(output_path="./output/reconciliation.xlsx")
    """

    def __init__(
        self,
        json_dir: Union[str, Path],
        excel_path: Union[str, Path],
        api_key: Optional[str] = None,
        endpoint: Optional[str] = None,
        deployment: Optional[str] = None,
        api_version: Optional[str] = None,
    ):
        self.json_dir = Path(json_dir)
        self.excel_path = Path(excel_path)

        if not self.json_dir.is_dir():
            raise FileNotFoundError(f"JSON directory not found: {self.json_dir}")
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        # Reuse the project's GPT extractor for LLM calls
        from src.extractors.gpt4_extractor import GPT4VisionExtractor

        self.extractor = GPT4VisionExtractor(
            api_key=api_key,
            endpoint=endpoint,
            deployment=deployment,
            api_version=api_version,
        )

        # Load Excel once
        logger.info(f"Loading Excel file: {self.excel_path}")
        self.df = pd.read_excel(self.excel_path, sheet_name="Export")
        logger.info(f"  Loaded {len(self.df)} rows, columns: {list(self.df.columns)}")

    def run(self, output_path: Union[str, Path] = "./output/reconciliation.xlsx") -> list[dict]:
        """
        Run the full reconciliation.

        Returns:
            List of dicts with keys: ses_number, sap_price, invoice_price, source_file
        """
        json_files = sorted(self.json_dir.glob("*.json"))
        if not json_files:
            logger.warning(f"No JSON files found in {self.json_dir}")
            return []

        logger.info(f"Found {len(json_files)} JSON files to process")
        records = []

        for jf in json_files:
            logger.info(f"\nProcessing: {jf.name}")

            raw_json = jf.read_text(encoding="utf-8")
            try:
                data = json.loads(raw_json)
            except json.JSONDecodeError as e:
                logger.error(f"  Skipping {jf.name}: invalid JSON ({e})")
                continue

            # Step 1: Extract SES Number
            ses_number = extract_ses_number(data, raw_json, self.extractor)
            if not ses_number:
                logger.warning(f"  Skipping {jf.name}: no SES Number found")
                continue

            # Step 2: Extract Document Total (Invoice Price)
            invoice_price = extract_document_total(data, raw_json, self.extractor)

            # Step 3: Lookup SAP Price
            try:
                sap_price = lookup_sap_price(self.df, ses_number)
            except (IndexError, KeyError) as e:
                logger.error(f"  Could not find SES/Spend columns in Excel: {e}")
                sap_price = 0.0

            records.append({
                "ses_number": ses_number,
                "sap_price": sap_price,
                "invoice_price": invoice_price if invoice_price is not None else 0.0,
                "source_file": jf.name,
            })

        # Write output
        if records:
            write_output_excel(records, output_path)
        else:
            logger.warning("No records to write.")

        # Summary
        logger.info("\n" + "=" * 60)
        logger.info("RECONCILIATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"JSON files processed: {len(json_files)}")
        logger.info(f"Records produced: {len(records)}")
        for r in records:
            diff = r["invoice_price"] - r["sap_price"]
            flag = " *** MISMATCH ***" if abs(diff) > 0.01 else ""
            logger.info(
                f"  SES {r['ses_number']}: SAP={r['sap_price']:.2f}  "
                f"Invoice={r['invoice_price']:.2f}  Diff={diff:.2f}{flag}"
            )
        logger.info("=" * 60)

        return records


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def cli_main():
    """Command-line entry point for the reconciliation tool."""
    parser = argparse.ArgumentParser(
        description="Reconcile extracted invoice JSONs against SAP Excel data",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example:
  python -m src.utils.reconciler \\
      --json-dir ./output/intermediate \\
      --excel ./input/sap_export.xlsx \\
      --output ./output/reconciliation.xlsx
        """,
    )
    # parser.add_argument("--json-dir", required=True, help="Directory containing extracted JSON files")
    # parser.add_argument("--excel", required=True, help="Path to SAP Excel file with SES Number and Total Spend columns")
    parser.add_argument("--output", "-o", default="./output/reconciliation.xlsx", help="Output Excel path")
    parser.add_argument("--api-key", help="Azure OpenAI API key")
    parser.add_argument("--endpoint", help="Azure OpenAI endpoint URL")
    parser.add_argument("--deployment", help="Azure OpenAI deployment name")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")

    args = parser.parse_args()
    args.json_dir = "pdf_table_extractor/input/JSONs"
    args.excel = "pdf_table_extractor/input/Mona & Dyno_01Jan24-28Feb26_PR,PO,SES,Inv.xlsx"
    from src.utils.logger import setup_logger
    setup_logger("pdf_extractor", level="DEBUG" if args.verbose else "INFO")

    try:
        reconciler = Reconciler(
            json_dir=args.json_dir,
            excel_path=args.excel,
            api_key=args.api_key,
            endpoint=args.endpoint,
            deployment=args.deployment,
        )
        reconciler.run(output_path=args.output)
    except Exception as e:
        print(f"\nError: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    cli_main()
